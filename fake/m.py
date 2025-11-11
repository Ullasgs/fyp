#!/usr/bin/env python3
# backend.py
# - loads Excel
# - serves /plants, /data, /recommendation
# - single /secret endpoint: GET shows page with Set & Reset buttons
# - POST {} sets persistent TDS offset (-200) and publishes adjusted payload
# - POST {"reset": true} resets offset to 0 and restores raw values (published & cached)

import os, json, time, threading, traceback
from flask import Flask, jsonify, request, make_response
from flask_cors import CORS
from openpyxl import load_workbook
import paho.mqtt.client as mqtt

# ---- CONFIG ----
BROKER = "172.20.10.10"
MQTT_PORT = 1883
TOPIC_SENSOR = "sensor/tds_ph"
TOPIC_RECOMMEND = "sensor/recommendation"
EXCEL_PATH = "Hydroponics_Plant_Range.xlsx"
FLASK_HOST = "0.0.0.0"
FLASK_PORT = 5000

# ---- GLOBALS (simple caches) ----
_latest_sensor_raw = None       # The last raw payload received from ESP32 (dict) and ts
_latest_sensor = None           # The last adjusted payload (dict) and ts (this is what /data returns as sensor)
_latest_recommendation = {}     # { payload: {...}, ts: <float> }
_plants_cache = {"ts": 0, "plants": {}}  # store dict keyed by name

# persistent offset applied to incoming TDS readings (in ppm).
_tds_offset = 0.0

app = Flask(__name__)
CORS(app)

# ---- HELPERS: Excel loader ----
def parse_range(cell):
    if cell is None:
        raise ValueError("empty")
    s = str(cell).strip()
    if "-" in s:
        a,b = [float(''.join(ch for ch in p if (ch.isdigit() or ch=='.' or ch=='-'))) for p in s.split("-")[:2]]
        return (min(a,b), max(a,b))
    nums = ''.join(ch for ch in s if (ch.isdigit() or ch=='.' or ch=='-'))
    if nums == "":
        raise ValueError("no numeric in '" + s + "'")
    v = float(nums)
    return (v, v)

def load_plants_excel(path):
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    wb = load_workbook(filename=path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return {}
    headers = [str(h).lower() if h is not None else "" for h in rows[0]]

    def idx(keys):
        for k in keys:
            for i,h in enumerate(headers):
                if k in h: return i
        return None

    i_name = idx(["plant","crop","name"])
    i_ph   = idx(["ph"])
    i_tds  = idx(["tds","ec"])
    i_temp = idx(["temp","temperature"])
    if None in (i_name, i_ph, i_tds, i_temp):
        raise ValueError("Excel headers missing required columns. Found: " + str(headers))

    plants = {}
    for r in rows[1:]:
        try:
            if r[i_name] is None: continue
            name = str(r[i_name]).strip()
            ph_min, ph_max = parse_range(r[i_ph])
            tds_min, tds_max = parse_range(r[i_tds])
            tmp_min, tmp_max = parse_range(r[i_temp])
            plants[name.lower()] = {
                "name": name,
                "ph": [ph_min, ph_max],
                "tds": [tds_min, tds_max],
                "temp": [tmp_min, tmp_max]
            }
        except Exception:
            print("WARNING: skipping row (parse error):", r)
            continue
    return plants

def refresh_plants_cache(force=False):
    try:
        mtime = os.path.getmtime(EXCEL_PATH)
    except Exception:
        mtime = 0
    if force or mtime != _plants_cache.get("ts", 0):
        try:
            plants = load_plants_excel(EXCEL_PATH)
            _plants_cache["plants"] = plants
            _plants_cache["ts"] = mtime
            print(f"[PLANTS] Loaded {len(plants)} plants from Excel (mtime={mtime})")
        except Exception as e:
            print("[PLANTS] Failed to load Excel:", e)
            traceback.print_exc()

# ---- MQTT callbacks ----
def on_connect(client, userdata, flags, rc):
    print("[MQTT] connected rc=", rc)
    client.subscribe(TOPIC_SENSOR)
    client.subscribe(TOPIC_RECOMMEND)
    print(f"[MQTT] subscribed to {TOPIC_SENSOR} and {TOPIC_RECOMMEND}")

def try_get_float(value):
    try:
        return float(value)
    except Exception:
        return None

def apply_tds_offset_to_payload(raw_payload):
    """
    Return a new dict where numeric 'tds' has had _tds_offset applied.
    This does NOT mutate raw_payload.
    """
    global _tds_offset
    new = dict(raw_payload)
    # prefer 'tds' key, else try alternates
    for k in ("tds", "tds_ppm", "TDS"):
        if k in raw_payload:
            v = try_get_float(raw_payload.get(k))
            if v is not None:
                new["tds"] = v + _tds_offset
                break
    # if no tds keys found, do nothing
    return new

def on_message(client, userdata, msg):
    """
    When an MQTT message arrives on TOPIC_SENSOR:
      - store the raw payload in _latest_sensor_raw
      - compute adjusted = apply_tds_offset_to_payload(raw)
      - store adjusted in _latest_sensor (this is what /data returns in 'sensor')
    """
    global _latest_sensor_raw, _latest_sensor, _latest_recommendation
    topic = msg.topic
    try:
        payload_raw = msg.payload.decode(errors='ignore')
        payload = json.loads(payload_raw)
    except Exception as e:
        print("[MQTT] received non-json on", topic, "payload:", msg.payload[:200], "err:", e)
        return

    now = time.time()
    if topic == TOPIC_SENSOR:
        _latest_sensor_raw = {"payload": payload, "ts": now}
        adjusted = apply_tds_offset_to_payload(payload)
        _latest_sensor = {"payload": adjusted, "ts": now}
        print(f"[MQTT] sensor RAW <- {payload}  -> cached ADJUSTED {adjusted}  (ts {now})")
    elif topic == TOPIC_RECOMMEND:
        _latest_recommendation = {"payload": payload, "ts": now}
        print(f"[MQTT] recommend <- {payload} (ts {now})")
    else:
        print("[MQTT] message on unexpected topic", topic)

# ---- MQTT runner thread ----
def mqtt_runner():
    client = mqtt.Client()
    client.on_connect = on_connect
    client.on_message = on_message
    try:
        client.connect(BROKER, MQTT_PORT, 60)
    except Exception as e:
        print("[MQTT] connect error:", e)
        return
    client.loop_forever()

# ---- FLASK endpoints ----
@app.route("/health")
def health():
    return jsonify({"ok": True, "broker": BROKER, "tds_offset": _tds_offset})

@app.route("/plants")
def plants():
    refresh_plants_cache(force=False)
    plants = _plants_cache.get("plants", {})
    out = []
    for nm, v in plants.items():
        out.append({
            "name": v["name"],
            "ph": v["ph"],
            "tds": v["tds"],
            "temp": v["temp"]
        })
    return jsonify({"plants": out})

@app.route("/data")
def data():
    """
    Returns:
      - sensor: the adjusted payload (what UI should display)
      - sensor_raw: the raw payload received from ESP32 (for reset/display/debug)
      - ts: timestamp for adjusted cache
      - tds_offset: current persistent offset
    """
    sensor = _latest_sensor["payload"] if _latest_sensor else None
    ts = _latest_sensor["ts"] if _latest_sensor else None
    sensor_raw = _latest_sensor_raw["payload"] if _latest_sensor_raw else None
    ts_raw = _latest_sensor_raw["ts"] if _latest_sensor_raw else None
    return jsonify({"sensor": sensor, "sensor_raw": sensor_raw, "ts": ts, "ts_raw": ts_raw, "tds_offset": _tds_offset})

@app.route("/recommendation")
def recommendation():
    if _latest_recommendation:
        return jsonify({"recommendation": _latest_recommendation["payload"], "ts": _latest_recommendation["ts"]})
    else:
        return jsonify({"recommendation": None, "ts": None})

# -------------------------
# SINGLE SECRET ENDPOINT: /secret
# - GET: returns small page with Set and Reset buttons
# - POST {} sets _tds_offset = -200 and publishes adjusted payload (based on raw)
# - POST {"reset": true} sets _tds_offset = 0 and restores cached payload to raw, publishes raw
# -------------------------
@app.route("/secret", methods=["GET", "POST"])
def secret():
    global _latest_sensor_raw, _latest_sensor, _tds_offset
    if request.method == "POST":
        body = {}
        if request.data:
            try:
                body = json.loads(request.data.decode('utf-8'))
            except Exception:
                body = {}
        now = time.time()

        if body.get("reset"):
            # Clear persistent offset and restore raw payload as the cached payload
            _tds_offset = 0.0
            if _latest_sensor_raw and isinstance(_latest_sensor_raw.get("payload"), dict):
                new_payload = dict(_latest_sensor_raw["payload"])
            else:
                new_payload = {"tds": None, "ph": None, "temperature": None}
            action = "reset"
        else:
            # Set persistent offset to -200 and compute adjusted from raw
            _tds_offset = -200.0
            if _latest_sensor_raw and isinstance(_latest_sensor_raw.get("payload"), dict):
                # compute adjusted = raw + offset
                raw = _latest_sensor_raw["payload"]
                # try to get raw numeric tds
                raw_val = try_get_float(raw.get("tds") or raw.get("tds_ppm") or raw.get("TDS"))
                if raw_val is None:
                    new_tds = _tds_offset
                else:
                    new_tds = raw_val + _tds_offset
                new_payload = dict(raw)
                new_payload["tds"] = new_tds
            else:
                new_payload = {"tds": _tds_offset, "ph": None, "temperature": None}
            action = "set"

        # update adjusted cache
        _latest_sensor = {"payload": new_payload, "ts": now}
        print(f"[SECRET] action={action} set persistent TDS offset to {_tds_offset}; updated cached payload -> {new_payload} at {now}")

        # publish to MQTT so subscribers see it immediately (best-effort)
        try:
            tmpc = mqtt.Client()
            tmpc.connect(BROKER, MQTT_PORT, 5)
            tmpc.loop_start()
            tmpc.publish(TOPIC_SENSOR, json.dumps(new_payload), qos=0)
            time.sleep(0.1)
            tmpc.loop_stop()
            tmpc.disconnect()
            print("[SECRET] published forced payload to MQTT topic", TOPIC_SENSOR)
        except Exception as e:
            print("[SECRET] mqtt publish failed:", e)

        return jsonify({"ok": True, "tds_offset": _tds_offset, "new_payload": new_payload, "ts": now, "action": action})

    # GET -> small HTML page with Set and Reset buttons
    html = """
    <!doctype html>
    <html>
      <head>
        <meta charset="utf-8">
        <title>Secret — TDS offset control</title>
        <style>
          body{font-family:Arial;display:flex;align-items:center;justify-content:center;height:100vh;background:#fafafa}
          .wrap{background:#fff;padding:24px;border-radius:10px;box-shadow:0 8px 30px rgba(0,0,0,0.06);text-align:center}
          button{padding:12px 18px;border-radius:8px;border:0;background:#e53e3e;color:white;font-weight:700;cursor:pointer;margin:6px}
          button.secondary{background:#2563eb}
          .ok{margin-top:12px;color:green}
        </style>
      </head>
      <body>
        <div class="wrap">
          <h2>Secret control</h2>
          <p>Set persistent <b>TDS offset = -200</b> or <b>Reset</b> to 0 (restore raw ESP32 values).</p>
          <div>
            <button id="setBtn">Set offset -200</button>
            <button id="resetBtn" class="secondary">Reset offset to 0</button>
          </div>
          <div id="result" class="ok"></div>
        </div>
        <script>
          async function doPost(payload) {
            const r = await fetch('/secret', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(payload)});
            return r.json();
          }
          document.getElementById('setBtn').addEventListener('click', async function(){
            const res = await doPost({});
            document.getElementById('result').innerText = 'OK — tds_offset=' + res.tds_offset + ' new_tds=' + (res.new_payload && res.new_payload.tds ? res.new_payload.tds : 'N/A');
          });
          document.getElementById('resetBtn').addEventListener('click', async function(){
            const res = await doPost({reset:true});
            document.getElementById('result').innerText = 'OK — reset, tds_offset=' + res.tds_offset + ' new_tds=' + (res.new_payload && res.new_payload.tds ? res.new_payload.tds : 'N/A');
          });
        </script>
      </body>
    </html>
    """
    resp = make_response(html)
    resp.headers['Content-Type'] = 'text/html'
    return resp

if __name__ == "__main__":
    refresh_plants_cache(force=True)
    t = threading.Thread(target=mqtt_runner, daemon=True)
    t.start()
    print(f"[BACKEND] Starting Flask on http://{FLASK_HOST}:{FLASK_PORT}")
    app.run(host=FLASK_HOST, port=FLASK_PORT)
