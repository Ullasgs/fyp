#!/usr/bin/env python3
# backend.py  (improved)
# - reloads plant Excel on demand
# - robust MQTT logging and caching
# - consistent JSON endpoints for frontend

import os, json, time, threading, traceback
from flask import Flask, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import paho.mqtt.client as mqtt

# ---- CONFIG ----
BROKER = "172.20.10.10"
MQTT_PORT = 1883
TOPIC_SENSOR = "sensor/tds_ph"
TOPIC_RECOMMEND = "sensor/recommendation"
EXCEL_PATH = "Hydroponics_Plant_Range.xlsx"
FLASK_HOST = "0.0.0.0"
FLASK_PORT = 5000

# ---- GLOBALS (thread-safe-ish simple caches) ----
_latest_sensor = {}          # { payload: {...}, ts: <float> }
_latest_recommendation = {}  # { payload: {...}, ts: <float> }
_plants_cache = {"ts": 0, "plants": {}}  # store dict keyed by name

app = Flask(__name__)
CORS(app)

# ---- HELPERS: Excel loader ----
def parse_range(cell):
    if cell is None:
        raise ValueError("empty")
    s = str(cell).strip()
    if "-" in s:
        a,b = [float(''.join(ch for ch in p if (ch.isdigit() or ch=='.' or ch=='-'))) for p in s.split("-")[:2]]
        return (min(a,b), max(a,b))
    nums = ''.join(ch for ch in s if (ch.isdigit() or ch=='.' or ch=='-'))
    if nums == "":
        raise ValueError("no numeric in '" + s + "'")
    v = float(nums)
    return (v, v)

def load_plants_excel(path):
    """Return dict {lower_name: {name, ph:[min,max], tds:[min,max], temp:[min,max]}}"""
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    wb = load_workbook(filename=path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return {}
    headers = [str(h).lower() if h is not None else "" for h in rows[0]]

    def idx(keys):
        for k in keys:
            for i,h in enumerate(headers):
                if k in h: return i
        return None

    i_name = idx(["plant","crop","name"])
    i_ph   = idx(["ph"])
    i_tds  = idx(["tds","ec"])
    i_temp = idx(["temp","temperature"])
    if None in (i_name, i_ph, i_tds, i_temp):
        raise ValueError("Excel headers missing required columns. Found: " + str(headers))

    plants = {}
    for r in rows[1:]:
        try:
            if r[i_name] is None: continue
            name = str(r[i_name]).strip()
            ph_min, ph_max = parse_range(r[i_ph])
            tds_min, tds_max = parse_range(r[i_tds])
            tmp_min, tmp_max = parse_range(r[i_temp])
            plants[name.lower()] = {
                "name": name,
                "ph": [ph_min, ph_max],
                "tds": [tds_min, tds_max],
                "temp": [tmp_min, tmp_max]
            }
        except Exception:
            # keep going but log the row parse failure
            print("WARNING: skipping row (parse error):", r)
            continue
    return plants

def refresh_plants_cache(force=False):
    """Reload plants if file changed or forced"""
    try:
        mtime = os.path.getmtime(EXCEL_PATH)
    except Exception:
        mtime = 0
    if force or mtime != _plants_cache.get("ts", 0):
        try:
            plants = load_plants_excel(EXCEL_PATH)
            _plants_cache["plants"] = plants
            _plants_cache["ts"] = mtime
            print(f"[PLANTS] Loaded {len(plants)} plants from Excel (mtime={mtime})")
        except Exception as e:
            print("[PLANTS] Failed to load Excel:", e)
            traceback.print_exc()

# ---- MQTT callbacks ----
def on_connect(client, userdata, flags, rc):
    print("[MQTT] connected rc=", rc)
    client.subscribe(TOPIC_SENSOR)
    client.subscribe(TOPIC_RECOMMEND)
    print(f"[MQTT] subscribed to {TOPIC_SENSOR} and {TOPIC_RECOMMEND}")

def on_message(client, userdata, msg):
    global _latest_sensor, _latest_recommendation
    topic = msg.topic
    try:
        payload_raw = msg.payload.decode(errors='ignore')
        payload = json.loads(payload_raw)
    except Exception as e:
        print("[MQTT] received non-json on", topic, "payload:", msg.payload[:200], "err:", e)
        return

    now = time.time()
    if topic == TOPIC_SENSOR:
        _latest_sensor = {"payload": payload, "ts": now}
        print(f"[MQTT] sensor <- {payload}  (ts {now})")
    elif topic == TOPIC_RECOMMEND:
        _latest_recommendation = {"payload": payload, "ts": now}
        print(f"[MQTT] recommend <- {payload} (ts {now})")
    else:
        print("[MQTT] message on unexpected topic", topic)

# ---- MQTT thread ----
def mqtt_runner():
    client = mqtt.Client()
    client.on_connect = on_connect
    client.on_message = on_message
    try:
        client.connect(BROKER, MQTT_PORT, 60)
    except Exception as e:
        print("[MQTT] connect error:", e)
        return
    client.loop_forever()

# ---- FLASK endpoints ----
@app.route("/health")
def health():
    return jsonify({"ok": True, "broker": BROKER})

@app.route("/plants")
def plants():
    # always try reload â€” keeps frontend up-to-date if file replaced
    refresh_plants_cache(force=False)
    plants = _plants_cache.get("plants", {})
    out = []
    for nm, v in plants.items():
        out.append({
            "name": v["name"],
            "ph": v["ph"],
            "tds": v["tds"],
            "temp": v["temp"]
        })
    return jsonify({"plants": out})

@app.route("/data")
def data():
    if _latest_sensor:
        return jsonify({"sensor": _latest_sensor["payload"], "ts": _latest_sensor["ts"]})
    else:
        return jsonify({"sensor": None, "ts": None})

@app.route("/recommendation")
def recommendation():
    if _latest_recommendation:
        return jsonify({"recommendation": _latest_recommendation["payload"], "ts": _latest_recommendation["ts"]})
    else:
        return jsonify({"recommendation": None, "ts": None})

if __name__ == "__main__":
    # initial load
    refresh_plants_cache(force=True)

    # start mqtt
    t = threading.Thread(target=mqtt_runner, daemon=True)
    t.start()

    print(f"[BACKEND] Starting Flask on http://{FLASK_HOST}:{FLASK_PORT}")
    app.run(host=FLASK_HOST, port=FLASK_PORT)
