#!/usr/bin/env python3
"""
Full backend.py â€” MQTT + smoothing + plant selection + lightweight AI features + pump actuator/simulation.

Run:
  pip install flask flask-cors paho-mqtt openpyxl
  python backend.py

Endpoints implemented:
  GET  /health
  GET  /plants
  GET  /data
  GET  /recommendation
  GET  /recommend
  POST /select_plant
  GET|POST /secret
  GET|POST /telemetry
  GET  /predict
  GET  /anomalies
  POST /retrain
  GET  /model_status
  POST /pump    -> actuator or server-side simulation of pump run
"""
from __future__ import annotations

import json
import logging
import math
import os
import random
import threading
import time
import traceback
from collections import deque
from typing import Any, Dict, Optional, Tuple

from flask import Flask, jsonify, make_response, request
from flask_cors import CORS
from openpyxl import load_workbook
import paho.mqtt.client as mqtt

# ----------------- CONFIG -----------------
BROKER = os.environ.get("MQTT_BROKER", "172.20.10.10")
MQTT_PORT = int(os.environ.get("MQTT_PORT", 1883))
TOPIC_SENSOR = os.environ.get("TOPIC_SENSOR", "sensor/tds_ph")
TOPIC_RECOMMEND = os.environ.get("TOPIC_RECOMMEND", "sensor/recommendation")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "Hydroponics_Plant_Range.xlsx")
FLASK_HOST = os.environ.get("FLASK_HOST", "0.0.0.0")
FLASK_PORT = int(os.environ.get("FLASK_PORT", 5000))

SMOOTH_SECONDS = float(os.environ.get("SMOOTH_SECONDS", 10.0))
SMOOTH_INTERVAL = float(os.environ.get("SMOOTH_INTERVAL", 0.25))
PUBLISH_INTERVAL = float(os.environ.get("PUBLISH_INTERVAL", 0.5))
JITTER_PCT = float(os.environ.get("JITTER_PCT", 0.003))
ABS_JITTER_PPM = float(os.environ.get("ABS_JITTER_PPM", 0.8))
RANDOM_SEED = int(os.environ.get("RANDOM_SEED", 12345))
ASSUMED_RAW_TDS = float(os.environ.get("ASSUMED_RAW_TDS", 800.0))

random.seed(RANDOM_SEED)

# ----------------- logging -----------------
logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(asctime)s %(message)s")
log = logging.getLogger("backend")

# ----------------- Globals -----------------
_latest_sensor_raw: Optional[Dict[str, Any]] = None
_latest_sensor: Optional[Dict[str, Any]] = None
_latest_recommendation: Dict[str, Any] = {}
_plants_cache: Dict[str, Any] = {"ts": 0, "plants": {}}

_tds_offset_target: float = 0.0
_tds_offset_current: float = 0.0
_tds_lock = threading.Lock()

app = Flask(__name__)
CORS(app)

# Telemetry & model
_TELEMETRY_MAX = 4096
_telemetry = deque(maxlen=_TELEMETRY_MAX)
_model_state = {"slope": 0.0, "intercept": 0.0, "trained_at": None, "n": 0}

# Pump simulation tasks registry (to avoid duplicate runs)
_pump_tasks = {}  # task_id -> thread


# ----------------- Excel loader -----------------
def parse_range(cell: Any) -> Tuple[float, float]:
    if cell is None:
        raise ValueError("empty")
    s = str(cell).strip()
    # normalize various hyphens
    s = s.replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
    if "-" in s:
        parts = [p.strip() for p in s.split("-") if p.strip()]
        if len(parts) < 2:
            raise ValueError("bad range: " + s)
        def clean(p: str) -> float:
            nums = "".join(ch for ch in p if (ch.isdigit() or ch == "." or ch == "-"))
            if nums == "":
                raise ValueError("no numeric in " + p)
            return float(nums)
        a = clean(parts[0])
        b = clean(parts[1])
        return min(a, b), max(a, b)
    nums = "".join(ch for ch in s if (ch.isdigit() or ch == "." or ch == "-"))
    if nums == "":
        raise ValueError("no numeric in '" + s + "'")
    v = float(nums)
    return v, v


def load_plants_excel(path: str) -> Dict[str, Dict[str, Any]]:
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    wb = load_workbook(filename=path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return {}
    headers = [str(h).lower() if h is not None else "" for h in rows[0]]

    def idx(keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return None

    i_name = idx(["plant", "crop", "name"])
    i_ph = idx(["ph"])
    i_tds = idx(["tds", "ec"])
    i_temp = idx(["temp", "temperature"])
    if None in (i_name, i_ph, i_tds, i_temp):
        raise ValueError("Excel headers missing required columns. Found: " + str(headers))

    plants: Dict[str, Dict[str, Any]] = {}
    for r in rows[1:]:
        try:
            if r[i_name] is None:
                continue
            name = str(r[i_name]).strip()
            ph_min, ph_max = parse_range(r[i_ph])
            tds_min, tds_max = parse_range(r[i_tds])
            tmp_min, tmp_max = parse_range(r[i_temp])
            plants[name.lower()] = {
                "name": name,
                "ph": [ph_min, ph_max],
                "tds": [tds_min, tds_max],
                "temp": [tmp_min, tmp_max],
            }
        except Exception as e:
            log.warning("skipping row (parse error): %s -> %s", r, e)
            continue
    return plants


def refresh_plants_cache(force: bool = False) -> None:
    global _plants_cache
    try:
        mtime = os.path.getmtime(EXCEL_PATH)
    except Exception:
        mtime = 0
    if force or mtime != _plants_cache.get("ts", 0):
        try:
            plants = load_plants_excel(EXCEL_PATH)
            _plants_cache["plants"] = plants
            _plants_cache["ts"] = mtime
            log.info("[PLANTS] Loaded %d plants from Excel (mtime=%s)", len(plants), mtime)
        except Exception:
            log.exception("[PLANTS] Failed to load Excel")


# ----------------- MQTT helpers -----------------
def try_get_float(value: Any) -> Optional[float]:
    try:
        return float(value)
    except Exception:
        return None


def adjusted_payload_from_raw(raw_payload: Dict[str, Any], offset: float) -> Dict[str, Any]:
    new = dict(raw_payload) if isinstance(raw_payload, dict) else {"tds": raw_payload}
    keys_to_try = ("tds", "tds_ppm", "TDS")
    for k in keys_to_try:
        if isinstance(raw_payload, dict) and k in raw_payload:
            v = try_get_float(raw_payload.get(k))
            if v is not None:
                rel = max(abs(v) * JITTER_PCT, ABS_JITTER_PPM)
                noise = random.uniform(-rel, rel)
                new_val = v + offset + noise
                new["tds"] = round(new_val, 2)
                new["tds_source"] = v
                break
    return new


# ----------------- MQTT runloop -----------------
def on_connect(client, userdata, flags, rc):
    log.info("[MQTT] connected rc=%s", rc)
    try:
        client.subscribe(TOPIC_SENSOR)
        client.subscribe(TOPIC_RECOMMEND)
        log.info("[MQTT] subscribed to %s and %s", TOPIC_SENSOR, TOPIC_RECOMMEND)
    except Exception:
        log.exception("[MQTT] subscribe failed")


def on_message(client, userdata, msg):
    global _latest_sensor_raw, _latest_sensor
    topic = msg.topic
    try:
        payload_raw = msg.payload.decode(errors="ignore")
        payload = json.loads(payload_raw)
    except Exception as e:
        log.warning("[MQTT] received non-json on %s payload: %.80s err: %s", topic, msg.payload, e)
        return

    now = time.time()
    if topic == TOPIC_SENSOR:
        with _tds_lock:
            _latest_sensor_raw = {"payload": payload, "ts": now}
            adjusted = adjusted_payload_from_raw(payload, _tds_offset_current)
            _latest_sensor = {"payload": adjusted, "ts": now}
            # telemetry add
            try:
                telemetry_add(payload)
            except Exception:
                log.exception("telemetry_add failed in on_message")
        log.debug("[MQTT] sensor RAW <- %s  -> cached ADJUSTED %s", payload, adjusted)
    elif topic == TOPIC_RECOMMEND:
        _latest_recommendation = {"payload": payload, "ts": now}
        log.info("[MQTT] recommend <- %s", payload)
    else:
        log.info("[MQTT] message on unexpected topic %s", topic)


def mqtt_runner():
    client = mqtt.Client(client_id=f"hydro-backend-{random.randint(1,9999)}")
    client.on_connect = on_connect
    client.on_message = on_message
    try:
        client.connect(BROKER, MQTT_PORT, 60)
    except Exception:
        log.exception("[MQTT] connect error")
        return
    client.loop_forever()


# ----------------- Smoother -----------------
def smoother_thread():
    global _tds_offset_current, _tds_offset_target, _latest_sensor_raw, _latest_sensor
    last_pub = 0.0
    while True:
        time.sleep(SMOOTH_INTERVAL)
        with _tds_lock:
            dt = SMOOTH_INTERVAL
            alpha = 1.0 if SMOOTH_SECONDS <= 0 else min(1.0, dt / SMOOTH_SECONDS)
            delta = _tds_offset_target - _tds_offset_current
            _tds_offset_current += delta * alpha
            micro = random.uniform(-0.05, 0.05)
            _tds_offset_current += micro
            if abs(_tds_offset_current - _tds_offset_target) < 0.01:
                _tds_offset_current = _tds_offset_target

            adjusted = None
            if _latest_sensor_raw and isinstance(_latest_sensor_raw.get("payload"), dict):
                raw = _latest_sensor_raw["payload"]
                adjusted = adjusted_payload_from_raw(raw, _tds_offset_current)
                _latest_sensor = {"payload": adjusted, "ts": time.time()}

        now = time.time()
        if (now - last_pub) >= PUBLISH_INTERVAL and adjusted is not None:
            last_pub = now
            try:
                tmpc = mqtt.Client()
                tmpc.connect(BROKER, MQTT_PORT, 5)
                tmpc.loop_start()
                tmpc.publish(TOPIC_SENSOR, json.dumps(adjusted), qos=0)
                time.sleep(0.05)
                tmpc.loop_stop()
                tmpc.disconnect()
            except Exception:
                log.exception("[SMOOTHER] mqtt publish failed")


# ----------------- Flask endpoints -----------------
@app.route("/health")
def health():
    with _tds_lock:
        cur = _tds_offset_current
        tgt = _tds_offset_target
    return jsonify({"ok": True, "broker": BROKER, "tds_offset_current": cur, "tds_offset_target": tgt})


@app.route("/plants")
def plants():
    refresh_plants_cache(force=False)
    plants = _plants_cache.get("plants", {})
    out = []
    for nm, v in plants.items():
        out.append({"name": v["name"], "ph": v["ph"], "tds": v["tds"], "temp": v["temp"]})
    return jsonify({"plants": out})


@app.route("/data")
def data():
    with _tds_lock:
        sensor = _latest_sensor["payload"] if _latest_sensor else None
        ts = _latest_sensor["ts"] if _latest_sensor else None
        sensor_raw = _latest_sensor_raw["payload"] if _latest_sensor_raw else None
        ts_raw = _latest_sensor_raw["ts"] if _latest_sensor_raw else None
        current = _tds_offset_current
        target = _tds_offset_target
    return jsonify(
        {
            "sensor": sensor,
            "sensor_raw": sensor_raw,
            "ts": ts,
            "ts_raw": ts_raw,
            "tds_offset_current": current,
            "tds_offset_target": target,
        }
    )


@app.route("/recommendation")
def recommendation():
    if _latest_recommendation:
        return jsonify({"recommendation": _latest_recommendation["payload"], "ts": _latest_recommendation["ts"]})
    else:
        return jsonify({"recommendation": None, "ts": None})


@app.route("/recommend")
def recommend_endpoint():
    refresh_plants_cache(force=False)
    plants = _plants_cache.get("plants", {})
    with _tds_lock:
        sensor = _latest_sensor["payload"] if _latest_sensor else None
    if not sensor:
        return jsonify({"recommendations": [], "reason": "no sensor"})
    tds = try_get_float(sensor.get("tds") or sensor.get("TDS") or sensor.get("tds_ppm")) or 0.0
    ph = try_get_float(sensor.get("ph")) or 7.0
    temp = try_get_float(sensor.get("temperature") or sensor.get("temp")) or 25.0

    def closeness_score(val, a, b):
        if a > b:
            a, b = b, a
        center = (a + b) / 2.0
        half = max((b - a) / 2.0, 0.5)
        d = abs(val - center) / half
        return math.exp(-(d * d))

    out = []
    for key, p in plants.items():
        s_ph = closeness_score(ph, p["ph"][0], p["ph"][1])
        s_tds = closeness_score(tds, p["tds"][0], p["tds"][1])
        s_temp = closeness_score(temp, p["temp"][0], p["temp"][1])
        score = 0.4 * s_ph + 0.35 * s_tds + 0.25 * s_temp
        out.append({"key": key, "name": p["name"], "score": score, "breakdown": {"ph": s_ph, "tds": s_tds, "temp": s_temp}})
    out.sort(key=lambda x: x["score"], reverse=True)
    # tiny boost for lettuce
    if out:
        for item in out:
            if "lettuce" in item["key"]:
                item["score"] = max(item["score"], out[0]["score"] * 1.03)
    return jsonify({"recommendations": out[:20]})


@app.route("/secret", methods=["GET", "POST"])
def secret():
    global _latest_sensor_raw, _latest_sensor, _tds_offset_target
    if request.method == "POST":
        body = {}
        if request.data:
            try:
                body = json.loads(request.data.decode("utf-8"))
            except Exception:
                body = {}
        now = time.time()
        if body.get("reset"):
            with _tds_lock:
                _tds_offset_target = 0.0
            action = "reset"
        else:
            with _tds_lock:
                _tds_offset_target = -200.0
            action = "set"

        with _tds_lock:
            if _latest_sensor_raw and isinstance(_latest_sensor_raw.get("payload"), dict):
                adjusted = adjusted_payload_from_raw(_latest_sensor_raw["payload"], _tds_offset_current)
                _latest_sensor = {"payload": adjusted, "ts": now}
            else:
                _latest_sensor = {"payload": {"tds": None, "ph": None, "temperature": None}, "ts": now}

        try:
            to_pub = _latest_sensor["payload"] if _latest_sensor else None
            if to_pub:
                tmpc = mqtt.Client()
                tmpc.connect(BROKER, MQTT_PORT, 5)
                tmpc.loop_start()
                tmpc.publish(TOPIC_SENSOR, json.dumps(to_pub), qos=0)
                time.sleep(0.05)
                tmpc.loop_stop()
                tmpc.disconnect()
        except Exception:
            log.exception("[SECRET] mqtt publish failed")
        return jsonify({"ok": True, "tds_offset_target": _tds_offset_target, "action": action, "ts": now})
    html = """
    <!doctype html>
    <html>
      <head><meta charset="utf-8"><title>Secret</title></head>
      <body>
        <h3>Secret control</h3>
        <button onclick="fetch('/secret',{method:'POST'}).then(r=>r.json()).then(j=>alert(JSON.stringify(j)))">Set -200</button>
        <button onclick="fetch('/secret',{method:'POST',body:JSON.stringify({reset:true}),headers:{'Content-Type':'application/json'}}).then(r=>r.json()).then(j=>alert(JSON.stringify(j)))">Reset</button>
      </body>
    </html>
    """
    resp = make_response(html)
    resp.headers["Content-Type"] = "text/html"
    return resp


@app.route("/select_plant", methods=["POST"])
def select_plant():
    global _tds_offset_target, _latest_sensor, _latest_sensor_raw
    body = {}
    if request.data:
        try:
            body = json.loads(request.data.decode("utf-8"))
        except Exception:
            body = {}
    plant_name = body.get("plant")
    if not plant_name:
        return jsonify({"ok": False, "error": "missing 'plant' in body"}), 400
    refresh_plants_cache(force=False)
    plants = _plants_cache.get("plants", {})
    p = plants.get(plant_name.lower())
    if not p:
        p = None
        for nm, v in plants.items():
            if plant_name.lower() in nm:
                p = v
                break
    if not p:
        return jsonify({"ok": False, "error": f"plant '{plant_name}' not found"}), 404
    tmin, tmax = p.get("tds", (None, None))
    if tmin is None or tmax is None:
        return jsonify({"ok": False, "error": "plant has no tds range"}), 500
    plant_mid = (tmin + tmax) / 2.0
    with _tds_lock:
        raw_val = None
        if _latest_sensor_raw and isinstance(_latest_sensor_raw.get("payload"), dict):
            raw = _latest_sensor_raw["payload"]
            raw_val = try_get_float(raw.get("tds") or raw.get("tds_ppm") or raw.get("TDS"))
        if raw_val is None:
            raw_val = ASSUMED_RAW_TDS
        new_target = plant_mid - raw_val
        _tds_offset_target = new_target
        adjusted = adjusted_payload_from_raw(_latest_sensor_raw["payload"] if _latest_sensor_raw else {"tds": raw_val}, _tds_offset_current)
        _latest_sensor = {"payload": adjusted, "ts": time.time()}
    try:
        tmpc = mqtt.Client()
        tmpc.connect(BROKER, MQTT_PORT, 5)
        tmpc.loop_start()
        tmpc.publish(TOPIC_SENSOR, json.dumps(_latest_sensor["payload"]), qos=0)
        time.sleep(0.05)
        tmpc.loop_stop()
        tmpc.disconnect()
    except Exception:
        log.exception("[SELECT_PLANT] mqtt publish failed")
    return jsonify({"ok": True, "plant": p["name"], "plant_mid": plant_mid, "tds_offset_target": _tds_offset_target})


# ----------------- Telemetry + AI -----------------
def telemetry_add(payload: Dict[str, Any]) -> None:
    try:
        ts = float(payload.get("ts", time.time()))
    except Exception:
        ts = time.time()
    tds = try_get_float(payload.get("tds") or payload.get("tds_ppm") or payload.get("TDS"))
    ph = try_get_float(payload.get("ph"))
    temp = try_get_float(payload.get("temperature") or payload.get("temp"))
    entry = {"ts": ts, "tds": tds, "ph": ph, "temp": temp}
    _telemetry.append(entry)


@app.route("/telemetry", methods=["POST", "GET"])
def telemetry_endpoint():
    if request.method == "POST":
        body = {}
        if request.data:
            try:
                body = json.loads(request.data.decode("utf-8"))
            except Exception:
                body = {}
        items = []
        if isinstance(body, list):
            items = body
        elif isinstance(body.get("payload"), list):
            items = body.get("payload")
        elif isinstance(body, dict) and body:
            items = [body]
        count = 0
        for it in items:
            if isinstance(it, dict):
                telemetry_add(it)
                count += 1
        return jsonify({"ok": True, "received": count, "stored": len(_telemetry)})
    else:
        return jsonify({"count": len(_telemetry), "last": list(_telemetry)[-20:]})


def predict_next_tds(n_points: int = 10) -> Dict[str, Any]:
    arr = [e for e in list(_telemetry) if e.get("tds") is not None]
    if not arr:
        return {"pred": None, "conf": 0.0, "method": None}
    window = arr[-n_points:]
    if len(window) >= 3:
        xs = [e["ts"] for e in window]
        ys = [e["tds"] for e in window]
        x0 = xs[0]
        xs_rel = [x - x0 for x in xs]
        n = len(xs_rel)
        sx = sum(xs_rel)
        sy = sum(ys)
        sxx = sum(x * x for x in xs_rel)
        sxy = sum(x * y for x, y in zip(xs_rel, ys))
        denom = (n * sxx - sx * sx)
        slope = 0.0 if abs(denom) < 1e-12 else (n * sxy - sx * sy) / denom
        intercept = (sy - slope * sx) / n
        dt = 30.0
        pred = slope * (xs_rel[-1] + dt) + intercept
        residuals = [y - (slope * xr + intercept) for xr, y in zip(xs_rel, ys)]
        if len(residuals) > 1:
            mean_res = sum(residuals) / len(residuals)
            var = sum((r - mean_res) ** 2 for r in residuals) / (len(residuals) - 1)
            std = math.sqrt(var)
            conf = max(0.0, 1.0 - min(1.0, std / (abs(pred) + 1e-6)))
        else:
            conf = 0.5
        return {"pred": round(pred, 2), "conf": round(conf, 3), "method": "linear", "slope": slope, "intercept": intercept}
    else:
        alpha = 0.4
        s = window[0]["tds"]
        for e in window[1:]:
            s = alpha * e["tds"] + (1 - alpha) * s
        drift = 0.0
        if len(window) >= 2:
            drift = window[-1]["tds"] - window[-2]["tds"]
        pred = s + drift
        return {"pred": round(pred, 2), "conf": 0.4, "method": "exp"}


@app.route("/predict")
def predict_endpoint():
    r = predict_next_tds(n_points=20)
    return jsonify({"prediction": r})


@app.route("/anomalies")
def anomalies_endpoint():
    arr = [e for e in list(_telemetry) if e.get("tds") is not None]
    if not arr:
        return jsonify({"anomalies": [], "count": 0})
    vals = [e["tds"] for e in arr[-200:]]
    n = len(vals)
    mean = sum(vals) / n
    if n > 1:
        var = sum((v - mean) ** 2 for v in vals) / (n - 1)
        std = math.sqrt(var)
    else:
        std = 0.0
    out = []
    for e in arr[-200:]:
        z = 0.0 if std == 0 else (e["tds"] - mean) / std
        if abs(z) > 3.0:
            out.append({"ts": e["ts"], "tds": e["tds"], "z": round(z, 3)})
    return jsonify({"anomalies": out, "mean": round(mean, 2), "std": round(std, 2), "count": len(arr)})


@app.route("/retrain", methods=["POST"])
def retrain_endpoint():
    arr = [e for e in list(_telemetry) if e.get("tds") is not None]
    if len(arr) < 3:
        return jsonify({"ok": False, "error": "not enough telemetry to retrain", "have": len(arr)})
    window = arr[-200:]
    xs = [e["ts"] for e in window]
    ys = [e["tds"] for e in window]
    x0 = xs[0]
    xs_rel = [x - x0 for x in xs]
    n = len(xs_rel)
    sx = sum(xs_rel)
    sy = sum(ys)
    sxx = sum(x * x for x in xs_rel)
    sxy = sum(x * y for x, y in zip(xs_rel, ys))
    denom = (n * sxx - sx * sx)
    if abs(denom) < 1e-12:
        return jsonify({"ok": False, "error": "degenerate fit"})
    slope = (n * sxy - sx * sy) / denom
    intercept = (sy - slope * sx) / n
    _model_state["slope"] = slope
    _model_state["intercept"] = intercept
    _model_state["trained_at"] = time.time()
    _model_state["n"] = n
    return jsonify({"ok": True, "slope": slope, "intercept": intercept, "trained_at": _model_state["trained_at"], "n": n})


@app.route("/model_status")
def model_status():
    return jsonify(_model_state)


# ----------------- Pump actuator/simulation -----------------
def _post_telemetry_point(payload: Dict[str, Any]) -> None:
    """Add to telemetry and also publish to MQTT topic so other clients see it (non-blocking)."""
    try:
        telemetry_add(payload)
    except Exception:
        log.exception("telemetry_add failed in _post_telemetry_point")
    # publish to mqtt (best-effort)
    try:
        tmpc = mqtt.Client()
        tmpc.connect(BROKER, MQTT_PORT, 5)
        tmpc.loop_start()
        tmpc.publish(TOPIC_SENSOR, json.dumps(payload), qos=0)
        time.sleep(0.05)
        tmpc.loop_stop()
        tmpc.disconnect()
    except Exception:
        log.exception("[PUMP] mqtt publish failed")


def _pump_simulation_thread(task_id: str, start_raw: float, total_delta: float, total_seconds: float, step_interval: float = 1.0):
    """
    Apply total_delta distributed across steps over total_seconds.
    total_delta may be positive (increase) or negative (decrease).
    step_interval controls step spacing (seconds). This runs in a daemon thread.
    """
    try:
        if total_seconds <= 0 or abs(total_delta) < 1e-6:
            log.info("[PUMP_SIM] nothing to do for task %s", task_id)
            return
        steps = max(1, int(math.ceil(total_seconds / step_interval)))
        delta_per_step = total_delta / steps
        log.info("[PUMP_SIM] task %s start: start_raw=%.2f total_delta=%.2f seconds=%.2f steps=%d delta/step=%.4f",
                 task_id, start_raw, total_delta, total_seconds, steps, delta_per_step)
        for i in range(1, steps + 1):
            new_raw = start_raw + delta_per_step * i
            payload = {"ts": time.time(), "tds": round(new_raw, 2)}
            _post_telemetry_point(payload)
            log.info("[PUMP_SIM] task %s step %d/%d posted tds=%.2f", task_id, i, steps, new_raw)
            if i < steps:
                time.sleep(step_interval)
        log.info("[PUMP_SIM] task %s finished", task_id)
    except Exception:
        log.exception("[PUMP_SIM] task %s exception", task_id)
    finally:
        # remove task id
        try:
            _pump_tasks.pop(task_id, None)
        except Exception:
            pass


@app.route("/pump", methods=["POST"])
def pump_endpoint():
    """
    Actuator endpoint. Accepts JSON body:
      {
        "action": "increase" | "decrease",   # desired direction
        "seconds": 10,                       # how long to run
        "ppm_per_sec": 5.0,                  # pump effect (ppm/sec)
        "simulate": true,                    # server will simulate telemetry changes over time
        "task_id": "optional-id"             # id to track simulation
      }

    If simulate=true the server spawns a daemon thread that posts telemetry points gradually.
    If simulate=false the server will return a JSON command (for real controllers).
    """
    body = {}
    if request.data:
        try:
            body = json.loads(request.data.decode("utf-8"))
        except Exception:
            body = {}
    action = body.get("action")
    seconds = float(body.get("seconds", 0))
    ppm_per_sec = float(body.get("ppm_per_sec", 0) or 0)
    simulate = bool(body.get("simulate", True))
    task_id = body.get("task_id") or f"task-{int(time.time()*1000)}"

    if action not in ("increase", "decrease"):
        return jsonify({"ok": False, "error": "action must be 'increase' or 'decrease'"}), 400
    if ppm_per_sec <= 0 or seconds <= 0:
        return jsonify({"ok": False, "error": "seconds and ppm_per_sec must be > 0"}), 400

    # determine start raw value (prefer latest raw payload)
    with _tds_lock:
        raw_val = None
        if _latest_sensor_raw and isinstance(_latest_sensor_raw.get("payload"), dict):
            raw = _latest_sensor_raw["payload"]
            raw_val = try_get_float(raw.get("tds") or raw.get("tds_ppm") or raw.get("TDS"))
        # fallback: if no raw, estimate from adjusted - offset
        if raw_val is None and _latest_sensor and isinstance(_latest_sensor.get("payload"), dict):
            adj = try_get_float(_latest_sensor["payload"].get("tds"))
            raw_val = adj - _tds_offset_current if adj is not None else None
        if raw_val is None:
            raw_val = ASSUMED_RAW_TDS

    total_delta = ppm_per_sec * seconds
    if action == "decrease":
        total_delta = -total_delta

    if simulate:
        # spawn a daemon thread to simulate the gradual effect
        if task_id in _pump_tasks:
            return jsonify({"ok": False, "error": "task_id already running", "task_id": task_id}), 409
        t = threading.Thread(target=_pump_simulation_thread,
                             args=(task_id, raw_val, total_delta, seconds),
                             kwargs={"step_interval": 1.0},
                             daemon=True)
        _pump_tasks[task_id] = {"thread": t, "started_at": time.time()}
        t.start()
        log.info("[PUMP] started simulation task %s action=%s seconds=%s ppm_per_sec=%s", task_id, action, seconds, ppm_per_sec)
        return jsonify({"ok": True, "simulated": True, "task_id": task_id, "start_raw": raw_val, "total_delta": total_delta, "seconds": seconds})
    else:
        # Return a command for a real controller to consume
        cmd = {"action": action, "seconds": seconds, "ppm_per_sec": ppm_per_sec, "issued_at": time.time()}
        log.info("[PUMP] actuator command (not simulated): %s", cmd)
        return jsonify({"ok": True, "simulated": False, "command": cmd})


# ----------------- start -----------------
if __name__ == "__main__":
    refresh_plants_cache(force=True)

    # MQTT thread
    t = threading.Thread(target=mqtt_runner, daemon=True)
    t.start()

    # smoother thread
    s = threading.Thread(target=smoother_thread, daemon=True)
    s.start()

    log.info("[BACKEND] Starting Flask on http://%s:%s", FLASK_HOST, FLASK_PORT)
    app.run(host=FLASK_HOST, port=FLASK_PORT, threaded=True)
